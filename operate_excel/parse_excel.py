# -*- coding: utf-8 -*-
# @Time : 2019/8/21 16:16
# @Author : wangmengmeng
from openpyxl import load_workbook
from openpyxl.styles import Font
from log.record_log import log


class ParseExcel:
    def __init__(self,filename):
        try:
            self.filename = filename
            self.wb = load_workbook(self.filename) # 打开现有的excel
        except FileNotFoundError as e:
            log.error("解析Excel文件{}失败\n{}".format(self.filename, e))

    def get_max_row_num(self,sheetname):
        """获取行数"""
        max_row_num = self.wb[sheetname].max_row
        return max_row_num

    def get_max_column_num(self,sheetname):
        """获取列数"""
        max_column = self.wb[sheetname].max_column
        return max_column

    def get_cell_value(self, sheet_name, row=None, column=None):
        """获取指定单元格的数据"""
        if isinstance(row, int) and isinstance(column, int):
            return self.wb[sheet_name].cell(row=row, column=column).value
        else:
            raise TypeError('row and column must be type int')

    def get_row_value(self,sheetname,row):
        """获取某一行的数据"""
        column_num = self.get_max_column_num(sheetname)
        row_value = []
        if isinstance(row, int):
            for column in range(1, column_num + 1):
                values_row = self.wb[sheetname].cell(row, column).value
                row_value.append(values_row)
            return row_value
        else:
            raise TypeError('row must be type int')

    def get_column_value(self,sheetname,column):
        """获取某一列的数据"""
        row_num = self.get_max_column_num(sheetname)
        column_value = []
        if isinstance(column, int):
            for row in range(1, row_num + 1):
                values_column = self.wb[sheetname].cell(row, column).value
                column_value.append(values_column)
            return column_value
        else:
            raise TypeError('column must be type int')

    def get_all_value(self,sheetname):
        """获取指定表单的所有数据（除去表头）"""














if __name__ == '__main__':
    pe = ParseExcel(DATA_PATH)


